import  openpyxl

class HandleExel:
    def __init__(self,filename,sheetname):
        """
        :param filename: 操作 excel文件名(路径)
        :param sheetname: 表单名
        """
        self.filename = filename
        self.sheetname = sheetname
    def read_data(self):
        """读取excel数据"""
        workbook = openpyxl.load_workbook(self.filename)
        sh = workbook[self.sheetname]
        res = list(sh.rows)
        # 获取第一行的表头
        title = [i.value for i in res[0]]
        # 遍历第一行之外的其它行
        cases = []
        for item in res[1:]:
            data = [i.value for i in item]
            dic = dict(zip(title, data))
            cases.append(dic)
        return cases
    def write_data(self,row,column,value):
        """
        数据写入的方法
        row:写入的行
        column:写入的列
        value:写入的值
        """
        # 加载工作簿对象
        workbook = openpyxl.load_workbook(self.filename)
        sh = workbook[self.sheetname]
        """写入数据"""
        sh.cell(row=row, column=column, value=value)
        workbook.save(self.filename)